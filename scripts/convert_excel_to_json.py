#!/usr/bin/env python
"""Convert an MSDS Excel workbook into website search JSON.

This script is intended for local use only. Keep real Excel files under
data/raw/ and write the converted output to data/msds.local.json, which is
ignored by Git.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SHEET_NAME = "통합종합_보기용"
DEFAULT_HEADER_ROW = 3

FIELD_ALIASES = {
    "no": ["NO", "No", "no", "번호"],
    "managementNo": ["관리번호", "관리 번호"],
    "productName": ["제품명", "제품 명"],
    "erpName": ["캠스 ERP 품명", "ERP 품명", "캠스ERP품명"],
    "msdsNo": ["MSDS번호", "MSDS 번호", "MSDS No", "MSDS NO"],
    "fileName": ["파일명", "파일 명", "PDF 파일명", "MSDS 파일명"],
    "category": ["용도분류", "용도 분류"],
    "recommendedUse": ["권고용도/사용용도", "권고용도", "사용용도", "권고 용도"],
    "supplier": ["제조사/공급업체", "제조사", "공급업체", "제조사 및 공급업체"],
    "emergencyContact": ["정보제공 및 긴급연락처", "긴급연락처", "정보제공"],
    "hazardClassification": ["주요 유해성 분류", "주요유해성분류", "유해성 분류"],
    "revisionDate": ["개정일", "개정 일자", "최종개정일"],
    "chemicalName": ["화학물질명", "화학 물질명", "물질명", "성분명"],
    "casNo": ["CAS No.", "CAS No", "CAS 번호", "CAS"],
    "content": ["함유량(%)", "함유량", "함량", "함량(%)"],
    "managementTarget": ["관리대상 유해물질", "관리대상유해물질", "관리대상 여부"],
    "workplaceMonitoringTarget": ["작업환경측정 대상", "작업환경측정대상", "작업환경측정"],
    "specialHealthCheckTarget": ["특수건강진단 대상", "특수건강진단대상", "특수건강진단"],
    "dangerousGoods": ["위험물 구분", "위험물구분", "위험물"],
    "ppeSummary": ["PPE 요약", "PPE", "보호구 요약", "개인보호구"],
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"[\s_\-/().]", "", text).lower()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def build_header_map(sheet, header_row: int) -> dict[str, int]:
    normalized_headers: dict[str, int] = {}
    for cell in sheet[header_row]:
        key = normalize_header(cell.value)
        if key and key not in normalized_headers:
            normalized_headers[key] = cell.column

    field_to_column: dict[str, int] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            column = normalized_headers.get(normalize_header(alias))
            if column:
                field_to_column[field] = column
                break
    return field_to_column


def get_row_value(row, field_to_column: dict[str, int], field: str) -> str:
    column = field_to_column.get(field)
    if not column:
        return ""
    return cell_text(row[column - 1].value)


def make_product_id(raw: dict[str, str], row_index: int) -> str:
    source = (
        raw.get("managementNo")
        or raw.get("no")
        or raw.get("msdsNo")
        or raw.get("fileName")
        or raw.get("productName")
        or f"row-{row_index}"
    )
    slug = re.sub(r"[^0-9a-zA-Z가-힣]+", "-", source).strip("-").lower()
    return slug or f"msds-{row_index}"


def pdf_path(file_name: str) -> str:
    return f"/pdf/{file_name}" if file_name else ""


def make_product(raw: dict[str, str], row_index: int) -> dict[str, Any]:
    return {
        "id": make_product_id(raw, row_index),
        "productName": raw.get("productName", ""),
        "erpName": raw.get("erpName", ""),
        "msdsNo": raw.get("msdsNo", ""),
        "fileName": raw.get("fileName", ""),
        "pdfPath": pdf_path(raw.get("fileName", "")),
        "category": raw.get("category", ""),
        "recommendedUse": raw.get("recommendedUse", ""),
        "supplier": raw.get("supplier", ""),
        "emergencyContact": raw.get("emergencyContact", ""),
        "hazardClassification": raw.get("hazardClassification", ""),
        "revisionDate": raw.get("revisionDate", ""),
        "dangerousGoods": raw.get("dangerousGoods", ""),
        "ppeSummary": raw.get("ppeSummary", ""),
        "ingredients": [],
        "siteLabel": "",
        "hazardBadge": "위험" if raw.get("hazardClassification") or raw.get("dangerousGoods") else "",
        "ghsPictograms": [],
        "hazardStatements": [],
        "precautionaryStatements": {
            "prevention": [],
            "response": [],
            "storage": [],
            "disposal": [],
        },
    }


def make_ingredient(raw: dict[str, str]) -> dict[str, str] | None:
    if not (raw.get("chemicalName") or raw.get("casNo") or raw.get("content")):
        return None
    return {
        "chemicalName": raw.get("chemicalName", ""),
        "casNo": raw.get("casNo", ""),
        "content": raw.get("content", ""),
        "managementTarget": raw.get("managementTarget", ""),
        "workplaceMonitoringTarget": raw.get("workplaceMonitoringTarget", ""),
        "specialHealthCheckTarget": raw.get("specialHealthCheckTarget", ""),
    }


def row_to_raw(row, field_to_column: dict[str, int]) -> dict[str, str]:
    return {field: get_row_value(row, field_to_column, field) for field in FIELD_ALIASES}


def is_product_row(raw: dict[str, str]) -> bool:
    return bool(raw.get("productName") or raw.get("fileName") or raw.get("managementNo"))


def convert(input_path: Path, output_path: Path, sheet_name: str, header_row: int) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise SystemExit(f"Sheet not found: {sheet_name}. Available sheets: {available}")

    sheet = workbook[sheet_name]
    field_to_column = build_header_map(sheet, header_row)
    products: list[dict[str, Any]] = []
    current_product: dict[str, Any] | None = None

    for row_index, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        raw = row_to_raw(row, field_to_column)
        if not any(raw.values()):
            continue

        if is_product_row(raw):
            current_product = make_product(raw, row_index)
            products.append(current_product)

        ingredient = make_ingredient(raw)
        if ingredient and current_product:
            current_product["ingredients"].append(ingredient)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(products, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return products


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert MSDS Excel workbook to website JSON.")
    parser.add_argument("--input", required=True, type=Path, help="Input Excel file path")
    parser.add_argument("--output", required=True, type=Path, help="Output JSON file path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Excel sheet name")
    parser.add_argument("--header-row", default=DEFAULT_HEADER_ROW, type=int, help="Header row number")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    products = convert(args.input, args.output, args.sheet, args.header_row)
    print(f"Converted {len(products)} products to {args.output}")


if __name__ == "__main__":
    main()
